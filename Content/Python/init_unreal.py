import unreal
import os
import openpyxl

@unreal.uclass()
class PMXlsxImporterPythonBridgeImpl(unreal.PMXlsxImporterPythonBridge):

    @unreal.ufunction(override = True)
    def read_worksheet_names(self, absolute_file_path):
        unreal.log("Reading xlsx file \"{0}\"".format(absolute_file_path))

        workbook = openpyxl.load_workbook(absolute_file_path, read_only=True, data_only = True)
        unreal.log("All sheet names: {0}".format(workbook.sheetnames))
        return workbook.sheetnames

    def parse_headers(self, row):
        headers = []
        for cell in row:
            headers.append(cell.value)
        return headers

    @unreal.ufunction(override = True)
    def read_worksheet(self, absolute_file_path, worksheet_name):
        workbook = openpyxl.load_workbook(absolute_file_path, read_only=True, data_only=True)
        worksheet = workbook[worksheet_name]

        results = []

        # with unreal.ScopedEditorTransaction("Import from xlsx") as transaction:
        # with unreal.ScopedSlowTask(worksheet.rows.length, "Importing from xlsx") as slow_task:
        # slow_task.make_dialog(True)
        headers = None
        row_index = 0
        for row in worksheet.rows:
            #if slow_task.should_cancel():
            #    break
            #slow_task.enter_progress_frame(i)

            if headers == None:
                headers = self.parse_headers(row)
                continue

            result = unreal.PMXlsxImporterPythonBridgeDataAssetInfo()

            column_index = 0
            for cell in row:
                # unreal.log("{0}: {1} {2}".format(column_index, headers[column_index], cell.value))
                # force keys and values to strings because unreal doesn't know how to convert other types automatically, and we want to pass a TMap<FString, FString> to unreal
                result.data[str(headers[column_index])] = str(cell.value) 
                column_index += 1

            result.asset_name = result.data['Name']
            results.append(result)

            row_index += 1

        return results
